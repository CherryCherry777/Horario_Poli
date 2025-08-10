#Quiero un programa que pueda leer el excel de la facultad y decirme qué horario y en qué aula hay x clase
#Adicionalmente, podré ver las aulas vacías en un horario especificado
#TODO servicio para guardar materias y poder consultar en cualquier momento
#TODO que busque aulas vacias en un tiempo determinado
import openpyxl
from openpyxl import Workbook
from colorama import Fore
import os
from tabulate import tabulate
from datetime import time
import unicodedata

print(Fore.GREEN + "Cargando..." + Fore.WHITE)
path = "horario.xlsx"
path2 = "horario_guardado.xlsx"

if (os.path.exists(path)):
  try:
    doc = openpyxl.load_workbook(path)
  except Exception as e:
    print("El horario no pudo ser cargado. Por favor verificar su existencia.")
    print("El archivo del horario debe ser de la facultad, nombrado horario.xlsx y estar en la misma carpeta que main.py")
else:
  print("El horario no pudo ser cargado. Por favor verificar su existencia.")
  print("El archivo del horario debe ser de la facultad, nombrado horario.xlsx y estar en la misma carpeta que main.py")

if (os.path.exists(path2)):
  try:
    print(Fore.GREEN + "Horario personalizado cargado exitosamente." + Fore.WHITE)
    doc2 = openpyxl.load_workbook(path2)
    hoja2 = doc2.active
  except Exception as e:
    print("El horario personalizado no pudo ser cargado.")
else:
  print(Fore.GREEN + "Creando archivo de horario personalizado..." + Fore.WHITE)
  #time.sleep(5)
  doc2 = Workbook()
  doc2.save(path2)
  hoja2 = doc2.active


#TODO try except every input so that there arent sudden errors



col_materia = 3 #columna de nombre de materia
col_seccion = 10 #columna de seccion de materia
col_prof_nombre = 14 #columna del nombre del docente
col_prof_apellido = 13 #columna del apellido del docente
col_prof_titulo = 12 #columna del titulo del docente
col_lunes = 36 #columna de las horas del lunes
col_martes = 38 #columna de las horas del martes
col_miercoles = 40 #columna de las horas del miercoles
col_jueves = 42 #columna de las horas del jueves
col_viernes = 44 #columna de las horas del viernes
col_sabado = 46 #columna de las horas del sabado

col_lunes_aula = 35
col_martes_aula = 37
col_miercoles_aula = 39
col_jueves_aula = 41
col_viernes_aula = 43
col_sabado_aula = 45
columnas = 47

global carreras
carreras = ["IAE","ICM","IEK","IEL","IEN","IIN","IMK","ISP","LCA","LCI","LCIk","LEL","LGH","TSE"]
global carreras_upper
carreras_upper = ["IAE","ICM","IEK","IEL","IEN","IIN","IMK","ISP","LCA","LCI","LCIK","LEL","LGH","TSE"]
"""
def refrescar_aulas():
  global aulas_a
  global aulas_b
  global aulas_c
  global aulas_e
  global aulas_f
  global aulas_h
  global aulas_i
  aulas_a = ["A50","A51","A52","A53","A54","A55","A56","A57","A58","A59"]
  aulas_b = ["B01","B02"]
  aulas_c = ["C01","C02","C03","C04"]
  aulas_e = ["E01","E02","E03","E04"]
  aulas_f = ["F05","F06","F07","F08","F09","F10","F11","F12","F13","F14","F15","F16","F17","F18","F25","F29","F30","F31","F33","F34","F35","F36","F37","F38","F39","F40"]
  aulas_h = ["H03","H04","H05","H06","H07","H08"]
  aulas_i = ["I01","I02","I03","I04","I05","I06","I07","I08"]
"""
def refrescar_aulas():
  global aulas
  aulas = [["A50","A51","A52","A53","A54","A55","A56","A57","A58","A59"],
  ["B01","B02"],
  ["C01","C02","C03","C04"],
  ["E01","E02","E03","E04"],
  ["F05","F06","F07","F08","F09","F10","F11","F12","F13","F14","F15","F16","F17","F18","F25","F29","F30","F31","F33","F34","F35","F36","F37","F38","F39","F40"],
  ["H03","H04","H05","H06","H07","H08"],
  ["I01","I02","I03","I04","I05","I06","I07","I08"]]
#Una celda vacia se presenta como "None"

#hoja = doc[nombreHoja]

def strip_accents(s):
   return ''.join(c for c in unicodedata.normalize('NFD', s)
                  if unicodedata.category(c) != 'Mn')

def crear_linea_horario(hoja, fila_materia):
  mat_aux = []
  #especificamente para imprimir UNA materia del horario
  #retorna un vector, para que imprima correctamente debe estar dentro de una matriz
  if(str(hoja.cell(row = fila_materia, column = col_lunes).value) == "None"):
    mat_aux.append("")
  else:
    mat_aux.append(hoja.cell(row=fila_materia,column = col_lunes_aula).value + " " + hoja.cell(row = fila_materia, column = col_lunes).value)

  if(str(hoja.cell(row = fila_materia, column = col_martes).value) == "None"):
    mat_aux.append("")
  else:
    mat_aux.append(hoja.cell(row=fila_materia,column = col_martes_aula).value + " " + hoja.cell(row = fila_materia, column = col_martes).value)

  if(str(hoja.cell(row = fila_materia, column = col_miercoles).value) == "None"):
    mat_aux.append("")
  else:
    mat_aux.append(hoja.cell(row=fila_materia,column = col_miercoles_aula).value + " " + hoja.cell(row = fila_materia, column = col_miercoles).value)

  if(str(hoja.cell(row = fila_materia, column = col_jueves).value) == "None"):
    mat_aux.append("")
  else:
    mat_aux.append(hoja.cell(row=fila_materia,column = col_jueves_aula).value + " " + hoja.cell(row = fila_materia, column = col_jueves).value)

  if(str(hoja.cell(row = fila_materia, column = col_viernes).value) == "None"):
    mat_aux.append("")
  else:
    mat_aux.append(hoja.cell(row=fila_materia,column = col_viernes_aula).value + " " + hoja.cell(row = fila_materia, column = col_viernes).value)

  if(str(hoja.cell(row = fila_materia, column = col_sabado).value) == "None"):
    mat_aux.append("")
  else:
    mat_aux.append(hoja.cell(row=fila_materia,column = col_sabado_aula).value + " " + hoja.cell(row = fila_materia, column = col_sabado).value)

  #print(tabulate(mat_aux,headers=["Lunes","Martes","Miércoles","Jueves","Viernes","Sábado"]))

  return mat_aux

def ver_horario():
  #esto es para imprimir el horario guardado
  mat_aux = []
  linea_aux = []
  for i in range(1,hoja2.max_row+1):
    if(hoja2.cell(row=i,column=col_materia).value is not None):
      linea_aux = crear_linea_horario(hoja2,i)
      linea_aux.insert(0,str(hoja2.cell(row=i,column=col_materia).value)+" "+str(hoja2.cell(row=i,column=col_seccion).value))
      mat_aux.append(linea_aux)

  #testing
  #print("mat_aux:",mat_aux)
  #print("linea_aux:",linea_aux)

  if(len(mat_aux) != 0):
    print(tabulate(mat_aux,headers=["Materia","Lunes","Martes","Miércoles","Jueves","Viernes","Sábado"]))
  else:
    print("\nNo hay materias guardadas.")

  print("\n\nPresione enter para continuar.")
  input()

  return

def eliminar_materia(fila_materia):
  hoja2.delete_rows(fila_materia)  
  doc2.save(path2)
  return

def eliminar_todas():
  for i in range(1,hoja2.max_row+1):
    hoja2.delete_rows(1)
    #print("Eliminada fila ",i)

  doc2.save(path2)
  return

def modificar_horario():

  for lp in range(100):
    if(hoja2.cell(row=1,column=1).value is not None):
      print("Elija la materia que desea eliminar")
      for i in range(1,hoja2.max_row+1):
        print(i,"-",hoja2.cell(row = i,column=col_materia).value,"-",hoja2.cell(row=i,column=col_seccion).value,"- Prof",hoja2.cell(row=i,column=col_prof_titulo).value,hoja2.cell(row=i,column=col_prof_nombre).value,hoja2.cell(row=i,column=col_prof_apellido).value)

      print(hoja2.max_row+1,"- Todas")
      print(hoja2.max_row+2,"- Volver al menu")
      try:
        seleccion = int(input())
      except Exception as e:
        print("Esa no es una seleccion valida.")
      else:
        if (0 < seleccion <= hoja2.max_row):
          print("Esta seguro? Escriba si para continuar, escriba otra cosa para cancelar.")
          try:
            conf = input().lower()
          except Exception as e:
            print(Fore.YELLOW+"No se ha eliminado la materia."+Fore.WHITE)
          else:
            if(conf == "si"):
              eliminar_materia(seleccion)
              print(Fore.RED+"Materia eliminada exitosamente."+Fore.WHITE)
              break
            else:
              print(Fore.YELLOW+"No se ha eliminado la materia."+Fore.WHITE)
              break
        elif(seleccion == hoja2.max_row+1):
          print("Esta seguro de querer eliminar todo? Escriba si para continuar, escriba otra cosa para cancelar.")
          try:
            conf = input().lower()
          except Exception as e:
            print(Fore.YELLOW+"No se ha borrado el horario."+Fore.WHITE)
          else:
            if(conf == "si"):
              eliminar_todas()
              print(Fore.RED+"Horario borrado exitosamente."+Fore.WHITE)
              break
            else:
              print(Fore.YELLOW+"No se ha borrado el horario."+Fore.WHITE)
              break
        elif(seleccion == hoja2.max_row+2):
          break
        else:
          print("Esa no es una seleccion valida.")
    else:
      print("\nNo hay materias guardadas.")
      print("\n\nPresione enter para continuar.")
      input()
      return
  return

def comparar_horas(inicio,fin,horario_inic,horario_fin):
  #devuelve 1 si se solapan, 0 si no
  if(inicio < horario_inic):
    if(fin<=horario_inic):
      #el bloque de tiempo es antes de la clase
      return 0
    elif(fin>horario_inic):
      #se intersectan
      return 1
    else:
      print(Fore.RED+"Error inesperado en la comparacion de horas.")
      return 2
  elif(inicio>horario_inic):
    if(inicio<horario_fin):
      #se intersectan
      return 1
    elif(inicio>=horario_fin):
      # bloque de tiempo comienza justo cuando termina la clase
      return 0
    else:
      print(Fore.RED+"Error inesperado en la comparacion de horas.")
      return 2
  elif(inicio == horario_inic):
    return 1
  else:
    print(Fore.RED+"Error inesperado en la comparacion de horas.")
    return 2


def separar_inicio_fin(horas):
  #devuelve un array de 2 donde tiene el inicio y fin de la asignatura
  #09:15 - 12:15
  #10:00 12:15
  aux1_hora = horas[0:2]
  aux1_min = horas[3:5]

  if(len(horas) == 11):
    aux2_hora = horas[6:8]
    aux2_min = horas[9:11]
  else:
    aux2_hora = horas[8:10]
    aux2_min = horas[11:13]
  
  try:
    inicio = time(int(aux1_hora),int(aux1_min))
    fin = time(int(aux2_hora),int(aux2_min))
  except Exception as e:
    print("Hubo un problema con la celda de contenido:", horas)
    input()
    quit()

  #print("Inicio:",inicio)
  #print("Fin:",fin)
  return [inicio,fin]

def buscar_aulas_vacias(dia,inicio,fin):
  #dia es la columna del dia que queremos revisar
  #inicio es el inicio del bloque de tiempo que estamos buscando
  #fin es el fin del bloque
  global aulas
  for carr in carreras:
    #print(Fore.YELLOW+"Carrera: "+Fore.WHITE+carr)
    hoja = doc[carr]
    for i in range(12,hoja.max_row+1):
      if(hoja.cell(row=i,column=dia).value is not None):
        hora = separar_inicio_fin(str(hoja.cell(row=i,column=dia).value))
        solapa = comparar_horas(inicio,fin,hora[0],hora[1])
        match solapa:
          case 1:
            #eliminar aula del array
            for bloque in aulas:
              if(hoja.cell(row=i,column=dia-1).value in bloque):
                bloque.remove(hoja.cell(row=i,column=dia-1).value)
          case 2:
            print("Ha ocurrido un error inesperado.")
            return

def verificar_duplicacion(hoja_p,fila_materia):
  #devuelve 1 si es duplicado
  #columna 1 es el codigo de clase, 5 es el codigo de carrera
  for i in range(1,hoja2.max_row + 1):
    if(hoja2.cell(row=i,column=1).value == hoja_p.cell(row=fila_materia,column=1).value and hoja2.cell(row=i,column=5).value == hoja_p.cell(row=fila_materia,column=5).value):
      return 1
    else:
      return 0
  return 0

def guardar_materia(hoja_p, fila_materia):
  #verificar_duplicacion(fila_materia)
  if(hoja2.max_row == 1 and hoja2.cell(row=1,column=1).value is not None):
    cant = hoja2.max_row
  elif(hoja2.max_row == 1 and hoja2.cell(row=1,column=1).value is None):
    cant = 0
  else:
    cant = hoja2.max_row

  for i in range (1, columnas + 1):
    hoja2.cell(row = cant + 1, column = i).value = hoja_p.cell(row = fila_materia, column = i).value
    
  doc2.save(path2)
  print("Materia añadida exitosamente.")
  print("\n\nPresione enter para continuar.")
  input()


def encontrar_aulas():
  #print("[green]Cargando...")
  #doc = openpyxl.load_workbook(path)
  print("\nIntroduzca el nombre de carrera (en siglas)")

  for lp in range(100):
    try:
      carrera = input().upper()
    except Exception as e:
      print("\nNo se pudo encontrar la carrera. Intente de nuevo.")
    else:
      if(carrera in carreras_upper):
        #esto es para asegurar que incluso si no se escribe con las mayusculas o minusculas correctas, igual funcione
        hoja = doc[carreras[carreras_upper.index(carrera)]]
        break
      else:
        print("\nNo se pudo encontrar la carrera. Intente de nuevo")

  # El test imprime la celda F12 que siempre tiene las siglas de la carrera
  # print("Test para ver si se abrio bien: ", hoja['F12'].value)
  for lp in range(100):
    print("\nIntroduzca el nombre de la materia (palabras claves, si no esta escrito exactamente no funciona.)")
    try:
      busqueda = input()
    except Exception as e:
      print("\nError. Por favor intente de nuevo.")

    resultados = [] #almacena los numeros de filas de resultados de la busqueda
    #print("Una celda vacia se presenta como: ", hoja.cell(row = filas + 1, column=1).value)

    for i in range(12, hoja.max_row):
      #test
      #print("Test: ",hoja.cell(row = i, column = 3).value)
      if (strip_accents(busqueda.lower()) in strip_accents(str(hoja.cell(row = i, column = col_materia).value).lower())):
        resultados.append(i)

    if (len(resultados) == 0):
      print("\nNo se encontro ninguna materia con ese nombre. Intente de nuevo.")
    else:
      break

  for i in range (len(resultados)):
    print(i+1,"-",hoja.cell(row=resultados[i],column=col_materia).value,"-",hoja.cell(row=resultados[i],column=col_seccion).value)
  
  print(len(resultados)+1,"-","Volver al menu")

  for lp in range(100):
    print("\nIntroduzca la clase de la cual quiere información")
    try:
      busqueda = int(input())
    except Exception as e:
      print("Error. Por favor intente de nuevo.")
    else:
      if(busqueda < 0 or busqueda > len(resultados) + 1):
        print("Esa no es una selección válida. Por favor intente de nuevo.")
    
  if(seleccion == len(resultados) + 1):
    print("\n")
    return
  
  seleccion = resultados[busqueda-1] #es la fila de la materia seleccionada

  print("Materia: ", hoja.cell(row = seleccion, column = col_materia).value, " - ", hoja.cell(row = seleccion, column = col_seccion).value)
  print("Profesor: ",hoja.cell(row = seleccion, column = col_prof_titulo).value,hoja.cell(row = seleccion, column= col_prof_nombre).value,hoja.cell(row = seleccion, column= col_prof_apellido).value)
  aux_dia = [[]]
  aux_dia[0] = crear_linea_horario(hoja,seleccion)
  print(tabulate(aux_dia,headers=["Lunes","Martes","Miércoles","Jueves","Viernes","Sábado"]))

  print(Fore.GREEN+"\nDesea guardar esta materia en su horario?"+Fore.WHITE)
  print("1- Si\n2- No\n")
  try:
    resp = int(input())
  except Exception as e:
    print("Por favor seleccione una de las opciones.")
  else:
    if (resp == 1):
      if(verificar_duplicacion(hoja,seleccion) == 0):
        guardar_materia(hoja, seleccion)
      else:
        print(Fore.RED+"Ya se ha registrado esa materia anteriormente."+Fore.WHITE)
    else:
      print(Fore.YELLOW+"No se ha guardado la materia."+Fore.WHITE)

  return


def encontrar_vacias():
  for lp in range(100):
    print("\nSeleccione el dia que quiere revisar.\n1-Lunes\n2-Martes\n3-Miercoles\n4-Jueves\n5-Viernes\n6-Sabado\n7-Volver")
    try:
      selec_dia = int(input())
    except Exception as e:
      print("\nPor favor ingrese un numero.")
    else:
      if (0 < selec_dia < 7):
        for lp in range(100):
          print("\nSeleccione el inicio del rango de tiempo que desea (formato HH:MM). Escriba salir para volver al menu.")
          try:
            selec_inicio = input()
            if(selec_inicio.lower() == "salir"):
              return
          except Exception as e:
            print("Error. Por favor intente otra vez.")
          else:
            
            try:
              #08:00
              #8:00
              if(len(selec_inicio) == 5):
                aux_1 = int(selec_inicio[0:2])
                aux_2 = int(selec_inicio[3:5])
                inicio = time(aux_1,aux_2)
              elif(len(selec_inicio) == 4):
                aux_1 = int(selec_inicio[0:1])
                aux_2 = int(selec_inicio[2:4])
                inicio = time(aux_1,aux_2)
              else:
                ("\nNo es un formato de hora aceptado.")
            except Exception as e:
              print("\nNo es un formato de hora aceptado.")
            else:
              for i in range(100):
                print("\nSeleccione el fin del rango de tiempo que desea (formato HH:MM). Escriba salir para volver al menu.")
                try:
                  selec_fin = input()
                  if(selec_inicio.lower() == "salir"):
                    return
                except Exception as e:
                  print("Error. Por favor intente otra vez.")
                try:
                  if(len(selec_fin) == 5):
                    aux_1 = int(selec_fin[0:2])
                    aux_2 = int(selec_fin[3:5])
                    fin = time(aux_1,aux_2)
                  elif(len(selec_fin) == 4):
                    aux_1 = int(selec_fin[0:1])
                    aux_2 = int(selec_fin[2:4])
                    fin = time(aux_1,aux_2)
                  else:
                    ("\nNo es un formato de hora aceptado.")
                except Exception as e:
                  print("\nNo es un formato de hora aceptado.")
                else:
                  if(inicio > fin):
                    print("La hora de fin no puede ser menor a la hora de inicio.")
                    return
                  match selec_dia:
                    case 1:
                      aux_dia = col_lunes
                    case 2:
                      aux_dia = col_martes
                    case 3:
                      aux_dia = col_miercoles
                    case 4:
                      aux_dia = col_jueves
                    case 5:
                      aux_dia = col_viernes
                    case 6:
                      aux_dia = col_sabado
                    case _:
                      print(Fore.RED+"\nSe ha producido un error.")
                      return
                  buscar_aulas_vacias(aux_dia,inicio,fin)
                  print("\nAulas vacias:")
                  print("Bloque A:",aulas[0])
                  print("Bloque B:",aulas[1])
                  print("Bloque C:",aulas[2])
                  print("Bloque E:",aulas[3])
                  print("Bloque F:",aulas[4])
                  print("Bloque H:",aulas[5])
                  print("Bloque I:",aulas[6])
                  break
            break  
      elif(selec_dia == 7):
        break
      else:
        print("\nEsa no es una seleccion valida.")
    #FOR TESTING
    #separar_inicio_fin("09:15 - 12:15")
    break
  refrescar_aulas()
  return

for lp in range(100):
  refrescar_aulas()
  #os.system('cls')
  print(Fore.CYAN+"\n\nBienvenido al programa de aulas!")
  print("\nQué desea hacer?")
  print("\n1-Ver horario y aula.\n2-Ver aulas vacías.\n3-Ver horario guardado\n4-Eliminar materia guardada\n5-Salir\n"+Fore.WHITE)
  try:
    accion = int(input())
  except Exception as e:
    print("Esa no es una opcion valida. Por favor intente otra vez.")

  match accion:
    case 1:
      encontrar_aulas()
    case 2:
      encontrar_vacias()
    case 3:
      ver_horario()
    case 4:
      modificar_horario()
    case 5:
      break
    case _:
      print("Esa no es una opción válida.")
  