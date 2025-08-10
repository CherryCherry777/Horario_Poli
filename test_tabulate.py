from tabulate import tabulate
import openpyxl
from openpyxl import Workbook
from colorama import Fore
import os
#

table = [["aeoifhwoeifw","b","c"]]

print(tabulate(table))

my_list = [1, 2, 3]
my_list.insert(0, 0)
print(my_list)

path = "testing.xlsx"
if (os.path.exists(path)):
  try:
    print(Fore.GREEN + "Horario personalizado cargado exitosamente." + Fore.WHITE)
    doc2 = openpyxl.load_workbook(path)
    hoja2 = doc2.active
  except Exception as e:
    print("El horario personalizado no pudo ser cargado.")
else:
  print(Fore.GREEN + "Creando archivo de horario personalizado..." + Fore.WHITE)
  #time.sleep(5)
  doc2 = Workbook()
  doc2.save(path)
  hoja2 = doc2.active

print("Para un archivo vacio hay",hoja2.max_row,"filas")
